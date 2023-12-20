import sqlite3
import openpyxl
from mapping import map
import os
import datetime
import pandas

with open('create_table.sql', 'r') as file:
    create_table_query = file.read()
with open('insert.sql', 'r') as file:
    insert_query = file.read()

workbook_path = 'Prelim25Cases_Database CC Case Data Collection_Tool 10.19.23.xlsx'
workbook = openpyxl.load_workbook(workbook_path, data_only=True)


db_dir= 'patient_cases.db'
if os.path.exists(db_dir):
    now = datetime.datetime.now()
    db_dir_old = 'old db as of '+ str(now)
    os.rename(db_dir, db_dir_old)
else:
    message = f"The file '{db_dir}' does not exist."

conn = sqlite3.connect('patient_cases.db')
cursor = conn.cursor()
cursor.execute(create_table_query)
def extract_data_from_sheet(sheet):
    return [sheet.cell(row=pos[0], column=pos[1]).value for field, pos in map.items()]

num_pages = int(input("Enter the number of pages you want to parse: "))

for sheet_name in workbook.sheetnames[:num_pages]:
    sheet = workbook[sheet_name]
    case_data = extract_data_from_sheet(sheet)
    placeholder_count = insert_query.count('?')
    print(f"Number of placeholders in insert_query: {placeholder_count}")
    print(f"Number of items in case_data: {len(case_data)}")
    print(case_data)
    cursor.execute(insert_query, case_data)
   
conn.commit()
conn.close()

