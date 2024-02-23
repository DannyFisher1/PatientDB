import sqlite3
import openpyxl
from mapping import map
import os
import datetime
import pandas as pd

with open('create_table.sql', 'r') as file:
    create_table_query = file.read()
with open('insert.sql', 'r') as file:
    insert_query = file.read()

workbook_path = 'New cases 0223.xlsx'
workbook = openpyxl.load_workbook(workbook_path, data_only=True)


db_dir= 'patient_cases.db'
if os.path.exists(db_dir):
    now = datetime.datetime.now()
    db_dir_old = 'old db as of '+ str(now)
    os.rename(db_dir, db_dir_old)
else:
    message = f"The file '{db_dir}' does not exist."

conn = sqlite3.connect('patient_cases.db')
cursor = conn.cursor()
cursor.execute(create_table_query)
def extract_data_from_sheet(sheet):
    return [sheet.cell(row=pos[0], column=pos[1]).value for field, pos in map.items()]
data_list = []

for sheet_name in workbook.sheetnames:
    if sheet_name[0] in ('0', '1', '2', '3', '4', '5', '6', '7', '8', '9'):
        sheet = workbook[sheet_name]
        case_data = extract_data_from_sheet(sheet)
        data_list.append(case_data)
        if len(case_data) != insert_query.count('?'):
            print(f"Mismatch between number of placeholders and data in sheet {sheet_name}. Skipping.")
            continue

        try:
            cursor.execute(insert_query, case_data)
        except Exception as e:
            print(f"Error inserting data for sheet {sheet_name}: {e}")
            continue
conn.commit()
conn.close()

df = pd.DataFrame(data_list)
print(df.head()) 