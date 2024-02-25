import sqlite3
import openpyxl
from mapping import map
import os
import datetime
import pandas as pd

def get_data(xl):
    print("Loading workbook...")
    workbook = openpyxl.load_workbook(xl, data_only=True)
    
    print("Reading SQL queries...")
    with open('database/create_table.sql', 'r') as file:
        create_table_query = file.read()
    with open('database/insert.sql', 'r') as file:
        insert_query = file.read()

    db_dir = 'database/patient_cases.db'
    if os.path.exists(db_dir):
        now = datetime.datetime.now()
        db_dir_old = 'old db as of ' + now.strftime("%Y-%m-%d %H:%M:%S")
        os.rename(db_dir, db_dir_old)
        print(f"Renamed existing database to: {db_dir_old}")
    else:
        print(f"The file '{db_dir}' does not exist, creating a new one.")

    conn = sqlite3.connect('patient_cases.db')
    cursor = conn.cursor()
    print("Creating database tables...")
    cursor.execute(create_table_query)

    def extract_data_from_sheet(sheet):
        return [sheet.cell(row=pos[0], column=pos[1]).value for field, pos in map.items()]
    print("Processing sheets...")
    for sheet_name in workbook.sheetnames:
        if sheet_name[0] in ('0', '1', '2', '3', '4', '5', '6', '7', '8', '9'):
                sheet = workbook[sheet_name]
                case_data = extract_data_from_sheet(sheet)
                if len(case_data) != insert_query.count('?'):
                    print(f"Mismatch in sheet {sheet_name}. Skipping.")
                    continue
                try:
                    cursor.execute(insert_query, case_data)
                    print(f"Data from sheet {sheet_name} inserted successfully.")
                except Exception as e:
                    print(f"Error in sheet {sheet_name}: {e}")
                    continue

    conn.commit()
    print("Data insertion complete.")

    print("Extracting and cleaning data...")
    query = "SELECT DISTINCT * FROM patient_cases WHERE first_bed_type IS NOT NULL AND first_bed_type <> 0;"
    df = pd.read_sql_query(query, conn)
    #fix numbers
    df['case_id'] = df['case_id'].apply(lambda x: x.split()[0] if isinstance(x, str) else x)

    conn.close()

    cleaned = df[[
        "case_id", "arrival_severity",
        "first_bed_type", "first_bed_hours",
        "second_bed_type", "second_bed_hours", 
        "third_bed_type", "third_bed_hours", 
        "fourth_bed_type", "fourth_bed_hours", 
        "fifth_bed_type", "fifth_bed_hours", 
        "sixth_bed_type", "sixth_bed_hours", 
        "seventh_bed_type", "seventh_bed_hours", 
        "eighth_bed_type", "eighth_bed_hours", 
        "ninth_bed_type", "ninth_bed_hours", 
        "tenth_bed_type", "tenth_bed_hours"
    ]]

    print("Exporting cleaned data to CSV...")
    cleaned.to_csv("csv/cleaned_patient_data.csv", index=False)
    df.to_csv("csv/patient_data.csv", index=False)
    print("All operations completed successfully.")
    return cleaned